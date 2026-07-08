"""Manual report exports for CSV, Excel, and PDF."""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

from . import database
from .calculations import minutes_to_hhmm, recalculate_range
from .config import DATA_DIR, ensure_data_dirs


EXPORT_DIR = DATA_DIR / "exports"


def export_period(
    conn,
    start_date: str,
    end_date: str,
    export_format: str,
    output_dir: str | Path | None = None,
) -> Path:
    """Export a date range as csv, xlsx, or pdf."""

    ensure_data_dirs()
    recalculate_range(conn, start_date, end_date)
    target_dir = Path(output_dir) if output_dir else EXPORT_DIR
    target_dir.mkdir(parents=True, exist_ok=True)
    fmt = export_format.lower().strip(".")
    rows = _build_rows(conn, start_date, end_date)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = target_dir / f"arbeitszeit_{start_date}_bis_{end_date}_{timestamp}.{fmt}"
    if fmt == "csv":
        _export_csv(path, rows)
    elif fmt in {"xlsx", "excel"}:
        path = path.with_suffix(".xlsx")
        _export_xlsx(path, rows)
    elif fmt == "pdf":
        _export_pdf(path, rows, start_date, end_date)
    else:
        raise ValueError("Exportformat muss csv, xlsx oder pdf sein.")
    return path


def _build_rows(conn, start_date: str, end_date: str) -> list[dict[str, Any]]:
    summaries = database.get_day_summaries_between(conn, start_date, end_date)
    notes = database.get_notes_between(conn, start_date, end_date)
    rows: list[dict[str, Any]] = []
    for summary in summaries:
        rows.append(
            {
                "Datum": summary["date"],
                "Kategorie": summary["day_category"],
                "Soll": minutes_to_hhmm(summary["target_minutes"]),
                "Ist": minutes_to_hhmm(summary["actual_minutes"]),
                "Pause": minutes_to_hhmm(summary["break_minutes"]),
                "Saldo": minutes_to_hhmm(summary["balance_minutes"]),
                "Standort": summary["location"] or "",
                "Notiz": notes.get(summary["date"], ""),
            }
        )
    return rows


def _export_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=_headers(), delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def _export_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("Excel-Export benoetigt openpyxl. Bitte requirements.txt installieren.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Arbeitszeit"
    sheet.append(_headers())
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([row[header] for header in _headers()])
    for index, header in enumerate(_headers(), start=1):
        width = max(len(header), *(len(str(row[header])) for row in rows)) + 2 if rows else len(header) + 2
        sheet.column_dimensions[get_column_letter(index)].width = min(width, 48)
    workbook.save(path)


def _export_pdf(path: Path, rows: list[dict[str, Any]], start_date: str, end_date: str) -> None:
    try:
        from fpdf import FPDF
    except ImportError as exc:
        raise RuntimeError("PDF-Export benoetigt fpdf2. Bitte requirements.txt installieren.") from exc

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 8, f"Arbeitszeit {start_date} bis {end_date}", ln=True)
    pdf.set_font("Helvetica", "B", 9)
    widths = [24, 30, 22, 22, 22, 22, 26, 120]
    for header, width in zip(_headers(), widths):
        pdf.cell(width, 7, header, border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 8)
    for row in rows:
        for header, width in zip(_headers(), widths):
            value = str(row[header]).encode("latin-1", "replace").decode("latin-1")
            pdf.cell(width, 6, value[:80], border=1)
        pdf.ln()
    pdf.output(str(path))


def _headers() -> list[str]:
    return ["Datum", "Kategorie", "Soll", "Ist", "Pause", "Saldo", "Standort", "Notiz"]

